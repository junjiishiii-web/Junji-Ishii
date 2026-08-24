"""
Leitura da SPEC (.xlsx) de um projeto URA/IVR da Claro.

Formato real observado (confirmado contra SPECs de producao, nao inventado):

Cada aba de estado segue o layout fixo:
    Nome do estado:      <nome>
    Estados anteriores:  <lista, pode continuar em linhas seguintes>
    Obs:                 <texto livre>
    <linha em branco>
    Nome do Prompt | Conteudo do Prompt | Contexto p/ Locutor | Observacoes | Marcacoes de B.I.
    <prompt 1>
    <prompt 2>
    ...
    Resultado | Vai para o estado... | Prompt que vai ouvir... | Observacoes | Marcacoes de B.I.
    <condicao 1 (pode ser hierarquica com prefixo ".." por nivel)>
    <condicao 2>
    ...

A aba `_Versionamento_` traz, por versao, a lista de estados alterados (coluna C)
e a descricao da alteracao (coluna D). A aba `VersionamentoBI` traz os
ScriptPoints novos publicados naquela versao.
"""
import datetime
import difflib
import re
import unicodedata

import openpyxl

# As SPECs "manuais" (template CRN/TLV antigo) usam colunas A-E lado a lado.
# SPECs exportadas pela ferramenta Rogue usam um layout bem mais largo (a
# marcacao real comeca la pela coluna I/K/L/N) com varias colunas de
# metadado extra a mais. Em vez de assumir uma posicao fixa, mapeamos cada
# campo pelo TEXTO do cabecalho (onde quer que ele esteja na linha) — isso
# funciona pros dois formatos, e pra variantes futuras, sem precisar saber a
# ferramenta de origem de antemao.
PROMPT_ID_ALIASES = ("nome do prompt", "prompt", "prompts", "id prompt", "id da fala")
PROMPT_TEXTO_ALIASES = ("conteúdo do prompt", "conteudo do prompt", "fala", "texto")
PROMPT_SP_ALIASES = ("marcações de b.i.", "marcacoes de b.i.", "anotação", "anotacao")

TRANSICAO_CONDICAO_ALIASES = ("resultado", "transição", "transicao", "condição", "condicao")
TRANSICAO_DESTINO_ALIASES = ("vai para o estado", "próximo estado", "proximo estado")
TRANSICAO_PROMPT_ALIASES = ("prompt que vai ouvir", "próximo prompt", "proximo prompt")
TRANSICAO_OBS_ALIASES = ("observações", "observacoes")
TRANSICAO_SP_ALIASES = ("marcações de b.i.", "marcacoes de b.i.", "anotação", "anotacao")

ESTADO_LABELS = ("nome do estado", "estado:")
ANTERIORES_LABELS = ("estados anteriores", "estado anterior")
OBS_LABELS = ("obs", "observações", "observacoes")

SP_INLINE_RE = re.compile(r"\b(spt|spp|ch|at|chr|et)\s*[:=]?\s*(\d{2,8})\b", re.IGNORECASE)
SP_BARE_RE = re.compile(r"\b(\d{4,8})\b")
# Formato real e dominante da Versionamento BI: "6846 - TLV_IDX_..." (codigo,
# hifen, descricao completa na MESMA celula) — ver secao 8 da skill.
CODE_WITH_DESC_RE = re.compile(r"^(\d{4,8})\s*[-–]\s*(.+)$")


def normalizar_chave(texto):
    if not texto:
        return ""
    txt = str(texto).strip().lower()
    txt = txt.replace(" ", "_")
    nfkd = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in nfkd if not unicodedata.combining(c))
    txt = re.sub(r"[^a-z0-9_]", "", txt)
    txt = re.sub(r"_+", "_", txt)
    return txt


def classificar_sp(texto):
    """Extrai (tipo, codigo) de uma marcacao de B.I. Ex.: 'ch = 62004' -> ('ch','62004')."""
    if not texto:
        return None
    m = SP_INLINE_RE.search(str(texto))
    if m:
        return (m.group(1).lower(), m.group(2))
    m = SP_BARE_RE.search(str(texto))
    if m:
        return ("spt", m.group(1))
    return None


def _normalizar_cabecalho(texto):
    return re.sub(r"\s+", " ", str(texto or "").strip().lower())


_SUFIXO_PONTUACAO_RE = re.compile(r"^[\s.…:]*$")


def _achar_coluna(normalizados, aliases):
    """Primeira coluna cujo texto normalizado bate com algum dos aliases —
    exato, ou por prefixo quando o que sobra depois do alias e so pontuacao/
    reticencias (ex.: 'vai para o estado...' bate com 'vai para o estado').
    NAO usa prefixo "solto": 'prompt que vai ouvir...' NAO pode bater com o
    alias 'prompt', porque o que sobra (' que vai ouvir...') tem palavras de
    verdade — isso classificaria errado o cabecalho de Transicao como se
    fosse o de Prompt."""
    for idx, val in enumerate(normalizados):
        if not val:
            continue
        for alias in aliases:
            if val == alias:
                return idx
            if val.startswith(alias) and _SUFIXO_PONTUACAO_RE.match(val[len(alias):]):
                return idx
    return None


def _mapear_cabecalho_prompt(normalizados):
    return {
        "id": _achar_coluna(normalizados, PROMPT_ID_ALIASES),
        "texto": _achar_coluna(normalizados, PROMPT_TEXTO_ALIASES),
        "sp": _achar_coluna(normalizados, PROMPT_SP_ALIASES),
    }


def _mapear_cabecalho_transicao(normalizados):
    return {
        "condicao": _achar_coluna(normalizados, TRANSICAO_CONDICAO_ALIASES),
        "destino": _achar_coluna(normalizados, TRANSICAO_DESTINO_ALIASES),
        "prox_prompt": _achar_coluna(normalizados, TRANSICAO_PROMPT_ALIASES),
        "obs": _achar_coluna(normalizados, TRANSICAO_OBS_ALIASES),
        "sp": _achar_coluna(normalizados, TRANSICAO_SP_ALIASES),
    }


def _valor(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def _materializar_linhas(sheet, max_col):
    """
    Le a aba inteira de uma vez via iter_rows (leitura sequencial). Em modo
    read_only, .cell(row, column) individual reprocessa o XML a cada chamada
    — para uma aba com ~80 linhas isso vira uma lentidao severa (varios
    segundos por aba). iter_rows() e a forma eficiente de ler tudo de uma vez,
    tanto em modo normal quanto em read_only.
    """
    # sheet.max_row pode vir None quando a aba nao declara <dimension> correta
    # no XML (comum em exports programaticos, ex.: ferramenta Rogue) mesmo
    # tendo dados reais. Nesse caso NAO se deve forcar max_row=1 (isso
    # truncava a aba inteira pra 1 linha!) — melhor omitir o limite e deixar
    # o iter_rows descobrir a extensao real pelas linhas de fato presentes.
    kwargs = {"min_row": 1, "min_col": 1, "max_col": max_col, "values_only": True}
    if sheet.max_row:
        kwargs["max_row"] = sheet.max_row
    return [list(row) for row in sheet.iter_rows(**kwargs)]


def parse_state_sheet(sheet):
    """
    Interpreta uma aba de estado. Os cabecalhos de Prompt/Transicao podem
    estar em qualquer coluna (template manual: colunas A-E; export do Rogue:
    colunas bem mais a direita, ~I/K/L/N, com varias colunas extras de
    metadado) — o mapeamento e feito pelo TEXTO do cabecalho, nao pela
    posicao. Abas vazias (sem nenhuma celula usada — openpyxl devolve
    max_row/max_column como None nesse caso) viram um estado sem
    prompts/transicoes, sem quebrar.
    """
    # Mesmo motivo do max_row: sheet.max_column vindo None nao significa aba
    # vazia — usa um teto generoso (30) nesse caso em vez de colapsar pra 6.
    max_col = max(6, min(sheet.max_column, 30)) if sheet.max_column else 30
    rows = _materializar_linhas(sheet, max_col)

    dados = {
        "nome": sheet.title,
        "estados_anteriores": [],
        "obs": "",
        "prompts": [],
        "transicoes": [],
    }

    section = None
    mapa_prompt = None
    mapa_transicao = None
    stack = {}
    i = 0
    total = len(rows)

    while i < total:
        row = rows[i]
        normalizados = [_normalizar_cabecalho(v) for v in row]
        col_a = str(row[0] or "").strip()
        col_a_low = col_a.lower().rstrip(":").strip()

        if any(col_a_low.startswith(lbl) for lbl in ESTADO_LABELS):
            valor = str(row[1] or "").strip()
            if valor:
                dados["nome"] = valor
            i += 1
            continue

        if any(col_a_low.startswith(lbl) for lbl in ANTERIORES_LABELS):
            vals = [str(v).strip() for v in row[1:] if v not in (None, "")]
            dados["estados_anteriores"].extend(vals)
            j = i + 1
            while j < total and not str(rows[j][0] or "").strip():
                more = [str(v).strip() for v in rows[j][1:] if v not in (None, "")]
                if not more:
                    break
                dados["estados_anteriores"].extend(more)
                j += 1
            i = j
            continue

        if any(col_a_low.startswith(lbl) for lbl in OBS_LABELS):
            obs_vals = [str(v).strip() for v in row[1:] if v not in (None, "")]
            dados["obs"] = " | ".join(obs_vals)
            i += 1
            continue

        # Deteccao de cabecalho de PROMPTS: exige achar a coluna de id +
        # pelo menos mais um campo (texto ou sp) na mesma linha, pra evitar
        # falso positivo de uma celula solta contendo "prompt".
        candidato_prompt = _mapear_cabecalho_prompt(normalizados)
        if candidato_prompt["id"] is not None and (candidato_prompt["texto"] is not None or candidato_prompt["sp"] is not None):
            section = "prompts"
            mapa_prompt = candidato_prompt
            stack = {}
            i += 1
            continue

        candidato_transicao = _mapear_cabecalho_transicao(normalizados)
        if candidato_transicao["condicao"] is not None and (candidato_transicao["destino"] is not None or candidato_transicao["sp"] is not None):
            section = "transicoes"
            mapa_transicao = candidato_transicao
            stack = {}
            i += 1
            continue

        if section == "prompts" and mapa_prompt:
            prompt_id = _valor(row, mapa_prompt["id"])
            texto = _valor(row, mapa_prompt["texto"])
            marcacao = _valor(row, mapa_prompt["sp"])
            eh_secao = not texto and prompt_id.upper() == prompt_id and len(prompt_id) > 3
            if prompt_id and texto and "exemplo" not in prompt_id.lower() and not eh_secao:
                dados["prompts"].append(
                    {
                        "id": prompt_id,
                        "texto": texto,
                        "sp": classificar_sp(marcacao),
                        "sp_raw": marcacao or None,
                    }
                )
            i += 1
            continue

        if section == "transicoes" and mapa_transicao:
            condicao_bruta = _valor(row, mapa_transicao["condicao"])
            if condicao_bruta:
                depth = 0
                label = condicao_bruta
                while label.startswith(".."):
                    depth += 1
                    label = label[2:].lstrip()
                label = label.strip()

                destino_raw = _valor(row, mapa_transicao["destino"])
                prox_prompt = _valor(row, mapa_transicao["prox_prompt"])
                obs_transicao = _valor(row, mapa_transicao["obs"])
                marcacao = _valor(row, mapa_transicao["sp"])

                stack[depth] = label
                for d in [d for d in stack if d > depth]:
                    del stack[d]

                if destino_raw and destino_raw.lower() not in ("nan",):
                    path = [stack[d] for d in sorted(stack) if d <= depth]
                    condicao = " > ".join(path) if path else label
                    dados["transicoes"].append(
                        {
                            "condicao": condicao,
                            "path": path,
                            "destino": destino_raw,
                            "prox_prompt": prox_prompt if prox_prompt.lower() not in ("nan", "") else "",
                            "obs": obs_transicao,
                            "sp": classificar_sp(marcacao),
                            "sp_raw": marcacao or None,
                        }
                    )
            i += 1
            continue

        i += 1

    dados["destinos_unicos"] = sorted({t["destino"] for t in dados["transicoes"] if t["destino"]})
    return dados


def _find_ivr_block_row(linhas, ivr_code=None):
    """Acha a ultima linha que contem 'IVR' (ou o codigo informado) na aba de versionamento.

    Specs reais escrevem o codigo do IVR com espacamento/pontuacao
    inconsistente entre blocos ("IVR-245786", "IVR- 245786", "IVR -245786"...).
    Comparar a string alvo inteira ("IVR-245786") como substring exata contra
    o texto da celula falha nesses casos (ex.: "IVR- 245786" tem um espaco
    a mais) — a busca nunca acha o bloco certo, cai no fallback de "sem
    bloco identificavel" e a aba INTEIRA (todas as versoes/projetos
    historicos do mesmo IVR, anos de mudancas nao relacionadas) e tratada
    como se fosse a versao atual, inflando os CTs gerados com centenas de
    SPs de outras versoes. Por isso comparamos so os DIGITOS do codigo,
    exigindo que a mesma celula tambem contenha "IVR" pra nao casar um
    numero solto por acaso.
    """
    alvo = (ivr_code or "").strip().upper()
    alvo_digitos = re.sub(r"\D", "", alvo)
    padrao_digitos = re.compile(r"(?<!\d)" + re.escape(alvo_digitos) + r"(?!\d)") if alvo_digitos else None
    melhor = -1
    for idx in range(len(linhas) - 1, -1, -1):
        r = idx + 1
        for val in linhas[idx]:
            val = str(val or "").strip()
            if not val:
                continue
            val_upper = val.upper()
            if alvo:
                if padrao_digitos and "IVR" in val_upper and padrao_digitos.search(val_upper):
                    return r
                if alvo in val_upper:
                    return r
            elif "IVR" in val_upper:
                melhor = r if melhor == -1 else melhor
        if melhor != -1 and not alvo:
            break
    return melhor


def parse_versionamento(wb, ivr_code=None, aba_candidatas=("_Versionamento_", "Versionamento")):
    aba = None
    for nome in aba_candidatas:
        if nome in wb.sheetnames:
            aba = wb[nome]
            break
    if not aba:
        return {"versao": None, "ivr": None, "responsavel": None, "estados_alterados": []}

    linhas = _materializar_linhas(aba, 10)
    row_ivr = _find_ivr_block_row(linhas, ivr_code)
    if row_ivr == -1:
        return {"versao": None, "ivr": None, "responsavel": None, "estados_alterados": []}

    linha_ivr = linhas[row_ivr - 1]
    versao = str(linha_ivr[0] or "").strip() or None
    ivr_texto = None
    for val in linha_ivr:
        val_str = str(val or "").strip()
        if "IVR" in val_str.upper() and ivr_texto is None:
            ivr_texto = val_str
    responsavel = str(linha_ivr[4] or "").strip() or None

    estados_alterados = []
    for idx in range(row_ivr, len(linhas)):
        linha = linhas[idx]
        estado_nome = str(linha[2] or "").strip()
        alteracao_desc = str(linha[3] or "").strip()
        if not estado_nome or estado_nome.lower() == "none":
            if not alteracao_desc:
                break
            continue
        if "ivr" in estado_nome.upper() or re.match(r"^v\.\d+$", estado_nome.lower()):
            break
        estados_alterados.append({"nome": estado_nome, "alteracao": alteracao_desc})

    return {
        "versao": versao,
        "ivr": ivr_texto,
        "responsavel": responsavel,
        "estados_alterados": estados_alterados,
    }


def _find_next_ivr_row(linhas, row_ivr):
    """row_ivr e 1-indexado (linha do cabecalho do bloco atual). Devolve o
    1-indexado da proxima linha que mencione 'IVR' (inicio do proximo
    bloco/versao), ou len(linhas)+1 se nao houver mais nenhum."""
    for idx in range(row_ivr, len(linhas)):
        for val in linhas[idx]:
            if val and "IVR" in str(val).upper():
                return idx + 1
    return len(linhas) + 1


def parse_versionamento_bi(wb, ivr_code=None, aba_candidatas=("VersionamentoBI", "_VersionamentoBI_", "Versionamento BI")):
    """
    A aba de Versionamento BI acumula o HISTORICO de todas as versoes/
    projetos, um bloco por IVR (igual a _Versionamento_) — nao e uma lista
    exclusiva do IVR atual. Sem escopar pelo bloco do IVR informado, SPs de
    OUTRAS versoes/projetos entram como se fossem novos desta, inflando os
    CTs gerados com ScriptPoints que nao fazem parte do escopo do projeto.
    """
    aba = None
    for nome in aba_candidatas:
        if nome in wb.sheetnames:
            aba = wb[nome]
            break
    if not aba:
        return []

    linhas = _materializar_linhas(aba, 10)
    row_ivr = _find_ivr_block_row(linhas, ivr_code)

    if row_ivr == -1:
        # Sem bloco de IVR identificavel (aba sem cabecalhos de versao, ou
        # codigo informado nao encontrado): mantem o comportamento antigo de
        # varrer a aba inteira, pra nao quebrar projetos greenfield sem
        # blocos de versionamento.
        linhas_do_bloco = linhas
    else:
        fim = _find_next_ivr_row(linhas, row_ivr)
        linhas_do_bloco = linhas[row_ivr - 1 : fim - 1]

    sps = []
    vistos = set()
    for valores in linhas_do_bloco:
        # Celulas de data/hora (ex.: quando a marcacao foi publicada) viram
        # string tipo "2026-07-10 00:00:00" via str() — isso bate por
        # acidente no MESMO formato regex de "codigo - descricao" ("2026"
        # seguido de hifen). Descarta qualquer valor que already seja
        # datetime/date/time NA ORIGEM, antes de virar string, pra nunca
        # cair nesse falso positivo.
        textos = [
            str(v).strip()
            for v in valores
            if v not in (None, "") and not isinstance(v, (datetime.date, datetime.datetime, datetime.time))
        ]
        if not textos:
            continue

        codigo = None
        descricao = None
        # Nome do estado costuma vir numa celula separada da mesma linha
        # (ex.: "SwitchIndicadorRentabilizacao"), sem nenhum digito — usamos
        # isso como heuristica pra distinguir do texto "codigo - descricao".
        estado = next((t for t in textos if not re.search(r"\d", t)), None)

        # Formato dominante real: "6846 - TLV_IDX_..." tudo na mesma celula.
        for t in textos:
            m = CODE_WITH_DESC_RE.match(t)
            if m:
                codigo, descricao = m.group(1), m.group(2).strip()
                break

        # Fallbacks: codigo puro numa celula, ou "spt=1234"/"ch = 1234"
        # embutido em texto (com a descricao na mesma celula, depois do
        # prefixo). Se so pegarmos o codigo e usarmos "o outro texto mais
        # longo" como descricao, o prefixo "ch = 1234 - " fica colado na
        # frente da descricao (ex.: "ch = 62004 - Chute_WhatsApp_Televendas"
        # em vez de so "Chute_WhatsApp_Televendas") — por isso extraimos a
        # descricao do que sobra APOS o match, nao do texto bruto.
        if not codigo:
            for t in textos:
                if re.fullmatch(r"\d{4,8}", t):
                    codigo = t
                    break
                m2 = SP_INLINE_RE.search(t)
                if m2:
                    codigo = m2.group(2)
                    resto = re.sub(r"^[\s\-–:]+", "", t[m2.end():])
                    if resto:
                        descricao = resto
                    break
            if codigo and not descricao:
                descricao = max((t for t in textos if t != codigo), key=len, default="")

        if not codigo or codigo in vistos:
            continue
        vistos.add(codigo)
        sps.append({"codigo": codigo, "descricao": descricao or "", "estado": estado})
    return sps


def _abrir_workbook(spec_path_or_bytes):
    """
    Abre a SPEC em modo read_only sempre que possivel. Isso evita um bug
    conhecido do openpyxl ao interpretar imagens/desenhos embutidos (ex.: logo
    da Claro colado na planilha) com atributos de fonte fora do range
    esperado (`ValueError: Max value is 52` em CharacterProperties.pitchFamily,
    que aparece encapsulado como "Unable to read workbook"). Em modo
    read_only o openpyxl nem tenta interpretar desenhos/imagens/graficos, so
    le celulas — que e tudo que este parser precisa.
    """
    if hasattr(spec_path_or_bytes, "seek"):
        spec_path_or_bytes.seek(0)
    try:
        return openpyxl.load_workbook(spec_path_or_bytes, data_only=True, read_only=True)
    except Exception:
        if hasattr(spec_path_or_bytes, "seek"):
            spec_path_or_bytes.seek(0)
        return openpyxl.load_workbook(spec_path_or_bytes, data_only=True)


# Nomes reservados de estado de fluxo no formato Rogue: "_Start_" (ponto de
# entrada, documentado na propria aba: 'Primeiro estado com o nome reservado
# "_Start_"') e "_End_" (destino de encerramento). Mesmo comecando com "_"
# como as abas de metadado (_SpecInfo_, _Doc_, _Calendar_ etc.), estes DOIS
# sao estados de verdade e precisam ficar em lista_abas/mapa_rotas — senao o
# BFS nunca acha o ponto de entrada e transicoes que apontam pra "_End_"
# ficam sem destino resolvido.
RESERVED_FLOW_STATES = {"_start_", "_end_"}


def _enriquecer_mapa_rotas_com_fuzzy(mapa_rotas, estados):
    """Specs reais tem erros de digitacao em ALGUMAS celulas de destino (ex.:
    uma transicao aponta pra 'BemVindo_TLVPosVendaClaroHDTV', mas a aba real
    se chama 'BemVindo_TLVPosVendasClaroHDTV' — falta um 's'). Isso quebra
    silenciosamente aquela UNICA aresta do fluxo, e o estado de destino fica
    permanentemente inalcancavel pro BFS mesmo estando estruturalmente
    conectado — o erro esta so no texto digitado, nao na topologia real do
    fluxo. Por isso, uma unica vez por leitura da SPEC (nao a cada chamada de
    BFS, que seria caro — uma spec grande tem milhares de transicoes),
    escaneamos todos os destinos citados e, pra qualquer um que nao resolva
    por igualdade exata, tentamos achar exatamente UM nome de aba real muito
    parecido (similaridade >= 0.92, sem ambiguidade) e registramos esse
    alias direto em mapa_rotas — resolver_destino nem percebe a diferenca.
    """
    chaves_reais = list(mapa_rotas.keys())
    vistos = set()
    for dados in estados.values():
        for t in dados.get("transicoes", []):
            destino_bruto = t.get("destino")
            if not destino_bruto:
                continue
            chave = normalizar_chave(destino_bruto)
            if not chave or len(chave) < 8 or chave in vistos or chave in mapa_rotas:
                continue
            vistos.add(chave)
            limpo = re.sub(r"^[\d]+[.\-_ ]+", "", chave)
            if limpo in mapa_rotas:
                continue
            proximos = difflib.get_close_matches(chave, chaves_reais, n=2, cutoff=0.92)
            if len(proximos) == 1:
                mapa_rotas[chave] = mapa_rotas[proximos[0]]


def read_spec(spec_path_or_bytes, ivr_code=None):
    """Le a SPEC inteira e devolve o modelo estrutural cru (sem regras de CT ainda)."""
    wb = _abrir_workbook(spec_path_or_bytes)
    lista_abas = [
        n
        for n in wb.sheetnames
        if (not n.startswith("_") or n.lower() in RESERVED_FLOW_STATES) and "versionamento" not in n.lower()
    ]
    mapa_rotas = {normalizar_chave(n): n for n in lista_abas}

    estados = {}
    for aba in lista_abas:
        estados[aba] = parse_state_sheet(wb[aba])

    _enriquecer_mapa_rotas_com_fuzzy(mapa_rotas, estados)

    versionamento = parse_versionamento(wb, ivr_code)
    versionamento_bi = parse_versionamento_bi(wb, ivr_code)

    return {
        "lista_abas": lista_abas,
        "estados": estados,
        "mapa_rotas": mapa_rotas,
        "versionamento": versionamento,
        "versionamento_bi": versionamento_bi,
    }


def resolver_destino(mapa_rotas, destino_bruto):
    if not destino_bruto:
        return None
    chave = normalizar_chave(destino_bruto)
    real = mapa_rotas.get(chave)
    if real:
        return real
    limpo = re.sub(r"^[\d]+[.\-_ ]+", "", chave)
    return mapa_rotas.get(limpo)

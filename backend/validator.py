"""
Validacao pos-geracao do Excel de cenarios: Gherkin valido, titulos <=255
chars, ultimo bloco e o REGRESSIVO, e presenca de SPs criticos. Porte da
secao 7 da skill de modelagem QA URA/BDD, adaptado para trabalhar com bytes
em memoria (sem depender de um caminho em disco).
"""
import io
import zipfile

import openpyxl
from lxml import etree

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

ABAS_CENARIOS_PADRAO = ["Cenarios", "Cenarios_MERGE", "Cenarios_ClaroEmpresas"]


def validar_xlsx(xlsx_bytes, abas_ct=None, sps_criticos=None):
    abas_ct = abas_ct or ABAS_CENARIOS_PADRAO
    sps_criticos = sps_criticos or []

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as z:
        if "xl/sharedStrings.xml" in z.namelist():
            ss_tree = etree.fromstring(z.read("xl/sharedStrings.xml"))
            shared = [
                "".join(t.text or "" for t in si.findall(f".//{{{NS}}}t"))
                for si in ss_tree.findall(f"{{{NS}}}si")
            ]
        else:
            shared = []

    relatorio = {"abas": [], "sps_criticos": [], "total_shared_strings": len(shared)}

    for aba in abas_ct:
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        ct_gherkin = {}
        for row in ws.iter_rows(min_col=1, max_col=3, values_only=True):
            v0 = str(row[0] or "").strip()
            vc = str(row[2] or "").strip()
            if v0.startswith("CT-") and vc:
                ct_gherkin[v0] = vc

        invalidos = [
            ct
            for ct, gh in ct_gherkin.items()
            if not (gh.lower().startswith("dado ") and "quando " in gh.lower() and "entao " in gh.lower().replace("então", "entao"))
        ]
        longos = [(ct, len(gh)) for ct, gh in ct_gherkin.items() if len(gh) > 255]

        ultimo_bloco = ""
        for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
            v = str(row[0] or "").strip()
            if v.lower().startswith("bloco") or v.lower() == "regressivo":
                ultimo_bloco = v

        relatorio["abas"].append(
            {
                "aba": aba,
                "total_cts": len(ct_gherkin),
                "gherkin_invalidos": invalidos,
                "gherkin_longos": longos,
                "ultimo_bloco": ultimo_bloco,
                "ultimo_bloco_ok": "regress" in ultimo_bloco.lower(),
            }
        )

    if sps_criticos:
        ss_str_lower = None
        for sp in sps_criticos:
            achou = any(sp in s for s in shared)
            relatorio["sps_criticos"].append({"sp": sp, "encontrado": achou})

    return relatorio


def resumo_textual(relatorio):
    linhas = []
    for aba in relatorio["abas"]:
        status_bloco = "OK" if aba["ultimo_bloco_ok"] else f"FALTA REGRESSIVO (último bloco: {aba['ultimo_bloco'] or '—'})"
        linhas.append(
            f"[{aba['aba']}] {aba['total_cts']} CTs | "
            f"Gherkin inválido: {len(aba['gherkin_invalidos'])} | "
            f"Gherkin >255 chars: {len(aba['gherkin_longos'])} | "
            f"Bloco final: {status_bloco}"
        )
    for sp in relatorio.get("sps_criticos", []):
        linhas.append(f"SP crítico {sp['sp']}: {'encontrado' if sp['encontrado'] else 'NÃO encontrado'}")
    return "\n".join(linhas)

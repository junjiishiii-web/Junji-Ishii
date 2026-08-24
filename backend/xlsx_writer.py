"""
Gera o Excel final (Capa, Planejamento, BI_Marcacoes, Legenda, Cenarios /
Cenarios_MERGE / Cenarios_ClaroEmpresas, Massa_Testes, Revisao_Necessaria),
no layout observado nas planilhas realmente entregues pela Mutant/Claro.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ct_rules import SP_PENDENTE_FILL

FONT_NAME = "Montserrat"
THIN = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
PRE_REQ_FILL = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
PENDENTE_FILL = PatternFill(start_color=SP_PENDENTE_FILL, end_color=SP_PENDENTE_FILL, fill_type="solid")


def _titulo(ws, texto, subtitulo=None):
    ws["B2"] = texto
    ws["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color="1F4E79")
    if subtitulo:
        ws["B3"] = subtitulo
        ws["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color="555555")


def _escrever_capa(wb, plano, eep_model, spec_model, ivr_code):
    ws = wb.active
    ws.title = "Capa"
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 70
    _titulo(ws, "MODELAGEM BDD/GHERKIN — QA SÊNIOR", "Mutant & Claro — Homologação de Canais URA/IVR")

    versionamento = spec_model.get("versionamento") or {}
    total_sp_pendente = sum(1 for c in plano["casos_teste"] if c["sp_pendente"])
    linhas = [
        ("Projeto:", f"{ivr_code or versionamento.get('ivr') or '—'} — {eep_model.get('nome', '—')}"),
        ("Canal/Operação:", "URA CRN/TLV — gerado automaticamente pelo Modeler Assistant"),
        ("Framework:", "Mutant / BDD Gherkin / XRAY"),
        ("Versão da SPEC:", versionamento.get("versao") or "—"),
        ("Total de Cenários:", f"{len(plano['casos_teste'])} CTs"),
        ("CTs com SP Pendente:", f"{total_sp_pendente} (revisar após publicação do BI)"),
        ("Chaves de Elegibilidade (EEP):", ", ".join(eep_model.get("chaves", [])) or "Nenhuma chave identificada"),
        ("Itens p/ Revisão Manual:", str(len(plano["revisao_necessaria"]))),
    ]
    for i, (rotulo, valor) in enumerate(linhas):
        row = i + 6
        ws.cell(row=row, column=2, value=rotulo).font = Font(name=FONT_NAME, size=10, bold=True)
        ws.cell(row=row, column=3, value=valor).font = Font(name=FONT_NAME, size=10)
    return ws


def _escrever_planejamento(wb, plano, eep_model):
    ws = wb.create_sheet("Planejamento")
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 90
    ws.append([None])
    ws["B2"] = "Seção"
    ws["C2"] = "Detalhamento Estratégico"
    for col in ("B", "C"):
        ws[f"{col}2"].font = Font(name=FONT_NAME, bold=True)
        ws[f"{col}2"].fill = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")

    blocos_resumo = ", ".join(b["titulo"] for b in plano["blocos"])
    linhas = [
        ("Escopo IN", ", ".join(eep_model.get("chaves", [])) or "Estados e ScriptPoints marcados na SPEC vigente."),
        ("Escopo OUT", "Indicadores/estados não citados na SPEC ou fora da cor de versão detectada."),
        (
            "Estratégia",
            f"{len(plano['casos_teste'])} CTs gerados automaticamente cobrindo os ScriptPoints "
            f"identificados no Versionamento BI. Perfis ANINÃO priorizados (L2).",
        ),
        ("Critérios Saída", "100% dos CTs executados e aprovados, sem SP PENDENTE em aberto."),
        ("Riscos", f"{len(plano['revisao_necessaria'])} item(ns) sinalizados para revisão manual (ver aba Revisão_Necessária)."),
        ("Resumo Blocos", blocos_resumo or "—"),
    ]
    for i, (rotulo, valor) in enumerate(linhas):
        row = i + 3
        ws.cell(row=row, column=2, value=rotulo).font = Font(name=FONT_NAME, bold=True)
        c = ws.cell(row=row, column=3, value=valor)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def _escrever_bi_marcacoes(wb, plano):
    ws = wb.create_sheet("BI_Marcacoes")
    headers = ["ScriptPoint", "Descrição da Marcação (Versionamento BI)", "CT Cobertura"]
    larguras = [14, 60, 20]
    for i, (h, w) in enumerate(zip(headers, larguras), start=1):
        ws.column_dimensions[chr(64 + i)].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    for i, linha in enumerate(plano["bi_marcacoes"], start=2):
        ws.cell(row=i, column=1, value=linha["codigo"])
        ws.cell(row=i, column=2, value=linha["descricao"])
        ws.cell(row=i, column=3, value=linha["ct"])


def _escrever_legenda(wb, plano):
    ws = wb.create_sheet("Legenda")
    headers = ["Bloco Visual", "Objetivo / Título do Bloco", "CTs Englobados", "Resumo SPs"]
    larguras = [14, 55, 18, 40]
    for i, (h, w) in enumerate(zip(headers, larguras), start=1):
        ws.column_dimensions[chr(64 + i)].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    for i, linha in enumerate(plano["legenda"], start=2):
        ws.cell(row=i, column=1, value=linha["bloco"])
        ws.cell(row=i, column=2, value=linha["titulo"])
        ws.cell(row=i, column=3, value=linha["cts"])
        ws.cell(row=i, column=4, value=linha["sps"])


def _escrever_massa_testes(wb, plano):
    ws = wb.create_sheet("Massa_Testes")
    headers = ["ID CT", "Perfil ANI", "Requisitos de Chave/API/Data", "Massa (Fornecer CPF/CNPJ)"]
    larguras = [10, 20, 55, 35]
    for i, (h, w) in enumerate(zip(headers, larguras), start=1):
        ws.column_dimensions[chr(64 + i)].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    for i, caso in enumerate(plano["casos_teste"], start=2):
        ws.cell(row=i, column=1, value=caso["ct_id"])
        ws.cell(row=i, column=2, value=f"{caso['perfil']} (CPF/CNPJ)")
        ws.cell(row=i, column=3, value="Especificados no Pré-Requisito do cenário")
        ws.cell(row=i, column=4, value="")


def _escrever_revisao_necessaria(wb, plano):
    ws = wb.create_sheet("Revisao_Necessaria")
    headers = ["Tipo", "Estado", "Detalhe"]
    larguras = [22, 24, 90]
    for i, (h, w) in enumerate(zip(headers, larguras), start=1):
        ws.column_dimensions[chr(64 + i)].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7B241C", end_color="7B241C", fill_type="solid")
    if not plano["revisao_necessaria"]:
        ws.cell(row=2, column=1, value="Nenhum item pendente de revisão manual — motor cobriu tudo com confiança.")
        return
    for i, item in enumerate(plano["revisao_necessaria"], start=2):
        ws.cell(row=i, column=1, value=item.get("tipo"))
        ws.cell(row=i, column=2, value=item.get("estado", "—"))
        c = ws.cell(row=i, column=3, value=item.get("detalhe"))
        c.alignment = Alignment(wrap_text=True, vertical="top")


def _escrever_cenarios(wb, aba_nome, casos_teste):
    ws = wb.create_sheet(aba_nome)
    larguras = {"A": 9, "B": 13, "C": 46, "D": 6, "E": 42, "F": 46, "G": 52}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    headers = ["ID CT", "Bloco", "Gherkin/Título BDD", "Step", "Ação do step", "Data / Estado / Prompt / Marcação", "Resultado Esperado (ScriptPoints)"]
    for col_idx, texto in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=texto)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    linha = 2
    from ct_rules import BLOCO_PALETTE, REGRESSIVO_COR

    for caso in casos_teste:
        cor = REGRESSIVO_COR if caso["regressivo"] else BLOCO_PALETTE[(caso["bloco"] - 1) % len(BLOCO_PALETTE)]
        start_row = linha

        header_fill = PatternFill(start_color=cor["header"], end_color=cor["header"], fill_type="solid")
        pendente_fill = PENDENTE_FILL if caso["sp_pendente"] else header_fill

        ws.cell(row=linha, column=1, value=caso["ct_id"]).fill = pendente_fill
        ws.cell(row=linha, column=1).font = Font(name=FONT_NAME, bold=True, color="000000" if caso["sp_pendente"] else "FFFFFF")
        ws.cell(row=linha, column=2, value=caso["bloco_nome"]).fill = header_fill
        ws.cell(row=linha, column=2).font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        ws.cell(row=linha, column=3, value=caso["gherkin"]).fill = PRE_REQ_FILL
        ws.cell(row=linha, column=3).alignment = Alignment(wrap_text=True, vertical="center")

        ws.cell(row=linha, column=4, value="0").fill = PRE_REQ_FILL
        pre_cell = ws.cell(row=linha, column=5, value="PRÉ-REQUISITO")
        pre_cell.fill = PRE_REQ_FILL
        pre_cell.font = Font(name=FONT_NAME, size=10, color="444444")
        cond_cell = ws.cell(row=linha, column=6, value=caso["pre_requisito"])
        cond_cell.fill = PRE_REQ_FILL
        cond_cell.font = Font(name=FONT_NAME, size=10, color="444444")
        cond_cell.alignment = Alignment(wrap_text=True, vertical="center")
        linha += 1

        for step in caso["steps"]:
            ws.cell(row=linha, column=4, value=step["step_num"])
            ws.cell(row=linha, column=5, value=step["acao"])
            estado_prompt = step["estado"]
            if step.get("prompt_id") not in (None, "—"):
                estado_prompt += f"\n{step['prompt_id']}"
            ws.cell(row=linha, column=6, value=estado_prompt)
            ws.cell(row=linha, column=7, value=step.get("resultado", step.get("prompt_texto", "—")))
            for col_idx in range(4, 8):
                cell = ws.cell(row=linha, column=col_idx)
                cell.font = Font(name=FONT_NAME, size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN
            linha += 1

        end_row = linha - 1
        for col in (1, 2, 3):
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=col).border = THIN
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

        linha += 1  # separador em branco

    return ws


def exportar_modelagem_para_xlsx(plano, eep_model, spec_model, ivr_code=None, aba_extra=None):
    """
    Monta o workbook final. `aba_extra`, quando fornecido, e uma tupla
    (nome_aba, casos_teste) para o modo dual SPEC (MERGE + ClaroEmpresas).
    """
    wb = Workbook()
    _escrever_capa(wb, plano, eep_model, spec_model, ivr_code)
    _escrever_planejamento(wb, plano, eep_model)
    _escrever_bi_marcacoes(wb, plano)
    _escrever_legenda(wb, plano)

    if aba_extra:
        _escrever_cenarios(wb, "Cenarios_MERGE", plano["casos_teste"])
        _escrever_cenarios(wb, "Cenarios_ClaroEmpresas", aba_extra[1])
    else:
        _escrever_cenarios(wb, "Cenarios", plano["casos_teste"])

    _escrever_massa_testes(wb, plano)
    _escrever_revisao_necessaria(wb, plano)
    return wb
